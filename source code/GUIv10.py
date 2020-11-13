from tkinter import *
import pandas as pd

from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from tkinter import ttk

root = Tk()

admins = {"admin":"password"}
employees={"employee":"password"}
supervisors = {"supervisor":"password"}
count=0
class Login(Frame):
    def __init__(self,master):
        master.title("User Login")
        master.geometry("500x500")
        for widget in master.winfo_children():
            widget.destroy()
        def validate(x,y):
            if x in admins:
                if admins[x]==y:
                    Admin(root)
                else:
                    invalid = Label(master,text="Invalid Login.",fg='red').place(relx=.5,rely=.8,anchor=CENTER)
            elif x in employees:
                if employees[x]==y:
                    Employee(root)
                else:
                    invalid = Label(master,text="Invalid Login.",fg='red').place(relx=.5,rely=.8,anchor=CENTER)
            elif x in supervisors:
                if supervisors[x]==y:
                    Supervisor(root)
                else:
                    invalid = Label(master,text="Invalid Login.",fg='red').place(relx=.5,rely=.8,anchor=CENTER)
            else:
                invalid = Label(master,text="Invalid Login.",fg='red').place(relx=.5,rely=.8,anchor=CENTER)
        username = StringVar()
        password=StringVar()
        usr_lbl = Label(master, text="Username: ")
        usr_lbl.place(relx=.3, rely=.4, anchor=CENTER)
        usr_text = Entry(master,textvariable=username, width=20)
        usr_text.place(relx=.5, rely=.4, anchor=CENTER)
        pwd_lbl = Label(master, text="Password: ")
        pwd_lbl.place(relx=.3, rely=.5, anchor=CENTER)
        pwd_text = Entry(master,textvariable=password, show="*", width=20)
        pwd_text.place(relx=.5, rely=.5, anchor=CENTER)
        submit = Button(master, text="Login",command=lambda: validate(username.get(),password.get()))
        submit.place(relx=.5, rely=.6, anchor=CENTER)
class AdminSupervisorDelete(Frame):
    def __init__(self,master):
        master.title("Delete Supervisor")
        master.geometry("500x500")
        def logOutBar():
            Login(root)
        for widget in master.winfo_children():
            widget.destroy()
        def deletes(x):
            supervisors.pop(x, None)
            Login(root)
        username2=StringVar()
        usr_lbl = Label(master, text="Name: ")
        usr_lbl.place(relx=.3, rely=.4, anchor=CENTER)
        usr_text = Entry(master,textvariable=username2, width=20)
        usr_text.place(relx=.5, rely=.4, anchor=CENTER)
        deleteSupervisor = Button(master, text="Delete Supervisor",command =lambda:deletes(username2.get()))
        deleteSupervisor.place(relx=.3, rely=.6, anchor=CENTER)
        view = Button(master,text = "View",command=viewSupervisors,width=30,height=5 )
        view.place(relx=.5, rely=.2, anchor=CENTER)
        add=Button(master,text = "Add",command=addSupervisors,width=30,height=5)
        add.place(relx=.5,rely=.5,anchor = CENTER)
        delete=Button(master,text="Delete",command=deleteSupervisors,width=30,height=5)
        delete.place(relx=.5,rely=.8,anchor=CENTER)
        #Employee menu bar
        menubar = Menu(master)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label = "Log Out",command=logOutBar)
        menubar.add_cascade(label="Log Out",menu=logoutMenu)
        master.config(menu=menubar)
class AdminSupervisorView(Frame):
    def __init__(self,master):
        master.title("View Supervisor")
        master.geometry("500x500")
        def logOutBar():
            Login(root)
        for widget in master.winfo_children():
            widget.destroy()
        view=Label(master,text=supervisors.keys()).pack()
        #Employee menu bar
        menubar = Menu(master)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label = "Log Out",command=logOutBar)
        menubar.add_cascade(label="Log Out",menu=logoutMenu)
        master.config(menu=menubar)
class AdminSupervisorAdd(Frame):
    def __init__(self,master):
        master.title("Add Supervisor")
        master.geometry("500x500")
        def logOutBar():
            Login(root)
        for widget in master.winfo_children():
            widget.destroy()
        def creates(x,y):
            supervisors[x]=y
            Login(root)
        username2=StringVar()
        password2=StringVar()
        usr_lbl = Label(master, text="Username: ")
        usr_lbl.place(relx=.3, rely=.4, anchor=CENTER)
        usr_text = Entry(master,textvariable=username2, width=20)
        usr_text.place(relx=.5, rely=.4, anchor=CENTER)
        pwd_lbl = Label(master, text="Password: ")
        pwd_lbl.place(relx=.3, rely=.5, anchor=CENTER)
        pwd_text = Entry(master,textvariable=password2, show="*", width=20)
        pwd_text.place(relx=.5, rely=.5, anchor=CENTER)
        createAccount = Button(master, text="Create Account",command =lambda: creates(username2.get(),password2.get()))
        createAccount.place(relx=.3, rely=.6, anchor=CENTER)
        #Employee menu bar
        menubar = Menu(master)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label = "Log Out",command=logOutBar)
        menubar.add_cascade(label="Log Out",menu=logoutMenu)
        master.config(menu=menubar)
class AdminSupervisor(Frame):
    def __init__(self,master):
        master.title("Edit Supervisor")
        master.geometry("500x500")
        def logOutBar():
            Login(root)
        def adminSupervisorView():
            AdminSupervisorView(root)
        def adminSupervisorAdd():
            AdminSupervisorAdd(root)
        def adminSupervisorDelete():
            AdminSupervisorDelete(root)
        for widget in master.winfo_children():
            widget.destroy()
        view = Button(master,text = "View",command=adminSupervisorView,width=30,height=5 )
        view.place(relx=.5, rely=.2, anchor=CENTER)
        add=Button(master,text = "Add",command=adminSupervisorAdd,width=30,height=5)
        add.place(relx=.5,rely=.5,anchor = CENTER)
        delete=Button(master,text="Delete",command=adminSupervisorDelete,width=30,height=5)
        delete.place(relx=.5,rely=.8,anchor=CENTER)
        #Employee menu bar
        menubar = Menu(master)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label = "Log Out",command=logOutBar)
        menubar.add_cascade(label="Log Out",menu=logoutMenu)
        master.config(menu=menubar)

        
class AdminEmployeeDelete(Frame):
    def __init__(self,master):
        master.title("Delete Employee")
        master.geometry("500x500")
        def logOutBar():
            Login(root)
        for widget in master.winfo_children():
            widget.destroy()
        def deletes(x):
            employees.pop(x, None)
            Login(root)
        username2=StringVar()
        usr_lbl = Label(master, text="Name: ")
        usr_lbl.place(relx=.3, rely=.4, anchor=CENTER)
        usr_text = Entry(master,textvariable=username2, width=20)
        usr_text.place(relx=.5, rely=.4, anchor=CENTER)
        deleteSupervisor = Button(master, text="Delete Employee",command =lambda:deletes(username2.get()))
        deleteSupervisor.place(relx=.3, rely=.6, anchor=CENTER)
        view = Button(master,text = "View",command=viewSupervisors,width=30,height=5 )
        view.place(relx=.5, rely=.2, anchor=CENTER)
        add=Button(master,text = "Add",command=addSupervisors,width=30,height=5)
        add.place(relx=.5,rely=.5,anchor = CENTER)
        delete=Button(master,text="Delete",command=deleteSupervisors,width=30,height=5)
        delete.place(relx=.5,rely=.8,anchor=CENTER)
        #Employee menu bar
        menubar = Menu(master)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label = "Log Out",command=logOutBar)
        menubar.add_cascade(label="Log Out",menu=logoutMenu)
        master.config(menu=menubar)
class AdminEmployeeView(Frame):
    def __init__(self,master):
        master.title("View Employee")
        master.geometry("500x500")
        def logOutBar():
            Login(root)
        for widget in master.winfo_children():
            widget.destroy()
        view=Label(master,text=employees.keys()).pack()
        #Employee menu bar
        menubar = Menu(master)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label = "Log Out",command=logOutBar)
        menubar.add_cascade(label="Log Out",menu=logoutMenu)
        master.config(menu=menubar)
class AdminEmployeeAdd(Frame):
    def __init__(self,master):
        master.title("Add Employee")
        master.geometry("500x500")
        def logOutBar():
            Login(root)
        for widget in master.winfo_children():
            widget.destroy()
        def creates(x,y):
            employees[x]=y
            Login(root)
        username2=StringVar()
        password2=StringVar()
        name = StringVar()
        address=StringVar()
        phoneNumber=StringVar()
        usr_lbl = Label(master, text="Username: ")
        usr_lbl.place(relx=.3, rely=.4, anchor=CENTER)
        usr_text = Entry(master,textvariable=username2, width=20)
        usr_text.place(relx=.5, rely=.4, anchor=CENTER)
        pwd_lbl = Label(master, text="Password: ")
        pwd_lbl.place(relx=.3, rely=.5, anchor=CENTER)
        pwd_text = Entry(master,textvariable=password2, show="*", width=20)
        pwd_text.place(relx=.5, rely=.5, anchor=CENTER)
        
        
        usr_lbl = Label(master, text="Name: ")
        usr_lbl.place(relx=.3, rely=.6, anchor=CENTER)
        usr_text = Entry(master,textvariable=name, width=20)
        usr_text.place(relx=.5, rely=.6, anchor=CENTER)
        
        usr_lbl = Label(master, text="Address: ")
        usr_lbl.place(relx=.3, rely=.7, anchor=CENTER)
        usr_text = Entry(master,textvariable=address, width=20)
        usr_text.place(relx=.5, rely=.7, anchor=CENTER)
        
        usr_lbl = Label(master, text="Phone Number: ")
        usr_lbl.place(relx=.3, rely=.8, anchor=CENTER)
        usr_text = Entry(master,textvariable=phoneNumber, width=20)
        usr_text.place(relx=.5, rely=.8, anchor=CENTER)
        
        
        createAccount = Button(master, text="Create Account",command =lambda: creates(username2.get(),password2.get()))
        createAccount.place(relx=.3, rely=.9, anchor=CENTER)
        #Employee menu bar
        menubar = Menu(master)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label = "Log Out",command=logOutBar)
        menubar.add_cascade(label="Log Out",menu=logoutMenu)
        master.config(menu=menubar)       
        
class AdminEmployee(Frame):
    def __init__(self,master):
        master.title("Edit Employee")
        master.geometry("500x500")
        def logOutBar():
            Login(root)
        def adminEmployeeView():
            AdminEmployeeView(root)
        def adminEmployeeAdd():
            AdminEmployeeAdd(root)
        def adminEmployeeDelete():
            AdminEmployeeDelete(root)
        for widget in master.winfo_children():
            widget.destroy()
        view = Button(master,text = "View",command=adminEmployeeView,width=30,height=5 )
        view.place(relx=.5, rely=.2, anchor=CENTER)
        add=Button(master,text = "Add",command=adminEmployeeAdd,width=30,height=5)
        add.place(relx=.5,rely=.5,anchor = CENTER)
        delete=Button(master,text="Delete",command=adminEmployeeDelete,width=30,height=5)
        delete.place(relx=.5,rely=.8,anchor=CENTER)
        #Employee menu bar
        menubar = Menu(master)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label = "Log Out",command=logOutBar)
        menubar.add_cascade(label="Log Out",menu=logoutMenu)
        master.config(menu=menubar)
class AdminBackup(Frame):
    def __init__(self,master):
        master.title("Edit Backup")
        master.geometry("500x500")
        def logOutBar():
            Login(root)
        for widget in master.winfo_children():
            widget.destroy()
        def createBackup():
            df=pd.read_excel("C:\\Hardware.xlsx",sheet_name="Hardware")
            df2=pd.read_excel("C:\\Hardware.xlsx",sheet_name="Electrical")
            df3=pd.read_excel("C:\\Hardware.xlsx",sheet_name="Plumbing")
            df4=pd.read_excel("C:\\Hardware.xlsx",sheet_name="Flooring")
            df5=pd.read_excel("C:\\Hardware.xlsx",sheet_name="Lumber")
            with pd.ExcelWriter('C:\\Users\CHR5S\Desktop\Backup.xlsx') as writer:  
                df.to_excel(writer,sheet_name="Hardware")
                df2.to_excel(writer,sheet_name="Electrical")
                df3.to_excel(writer,sheet_name="Plumbing")
                df4.to_excel(writer,sheet_name="Flooring")
                df5.to_excel(writer,sheet_name="Lumber")
                backupCreated=Label(master,text="Backup Created").pack()
        def restoreBackup():
            df=pd.read_excel("C:\\Users\CHR5S\Desktop\Backup.xlsx",sheet_name="Hardware")
            df2=pd.read_excel("C:\\Users\CHR5S\Desktop\Backup.xlsx",sheet_name="Electrical")
            df3=pd.read_excel("C:\\Users\CHR5S\Desktop\Backup.xlsx",sheet_name="Plumbing")
            df4=pd.read_excel("C:\\Users\CHR5S\Desktop\Backup.xlsx",sheet_name="Flooring")
            df5=pd.read_excel("C:\\Users\CHR5S\Desktop\Backup.xlsx",sheet_name="Lumber")
            with pd.ExcelWriter('C:\\Users\CHR5S\Desktop\Hardware.xlsx') as writer:  
                df.to_excel(writer,sheet_name="Hardware")
                df2.to_excel(writer,sheet_name="Electrical")
                df3.to_excel(writer,sheet_name="Plumbing")
                df4.to_excel(writer,sheet_name="Flooring")
                df5.to_excel(writer,sheet_name="Lumber")
                backupCreated=Label(master,text="Restored from backup").pack()
        view = Button(master,text = "Create",command=createBackup,width=30,height=5 )
        view.place(relx=.5, rely=.2, anchor=CENTER)
        add=Button(master,text = "Restore",command=restoreBackup,width=30,height=5)
        add.place(relx=.5,rely=.5,anchor = CENTER)
        #Employee menu bar
        menubar = Menu(master)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label = "Log Out",command=logOutBar)
        menubar.add_cascade(label="Log Out",menu=logoutMenu)
        master.config(menu=menubar)
class AdminInventory(Frame):
    def __init__(self,master):
        master.title("Edit Inventory")
        master.geometry("500x500")
        def logOutBar():
            Login(root)
        def viewInventory():
            Departments(root)
        for widget in master.winfo_children():
            widget.destroy()
        view = Button(master,text = "View",command=viewInventory,width=30,height=5 )
        view.place(relx=.5, rely=.2, anchor=CENTER)
        add=Button(master,text = "Add",width=30,height=5)
        add.place(relx=.5,rely=.5,anchor = CENTER)
        delete=Button(master,text="Delete",width=30,height=5)
        delete.place(relx=.5,rely=.8,anchor=CENTER)
        #Employee menu bar
        menubar = Menu(master)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label = "Log Out",command=logOutBar)
        menubar.add_cascade(label="Log Out",menu=logoutMenu)
        master.config(menu=menubar)
class Admin(Frame):
    def __init__(self,master):
        master.title("Admin")
        master.geometry("500x500")
        for widget in master.winfo_children():
            widget.destroy()
        #Admin functions
        def supervisorBar():
            AdminSupervisor(root)
        def backupBar():
            AdminBackup(root)
        def employeeBar():
            AdminEmployee(root)
        def inventoryBar():
            AdminInventory(root)
        def logOutBar():
            Login(root)
        def employee():
            Employee(root)
        def departments():
            Departments(root)
        def activityLog():
            Activity_Log(root)
        #Admin menu bar
        menubar = Menu(master)
        supervisorMenu = Menu(menubar, tearoff=0)
        supervisorMenu.add_command(label = "Supervisor",command=supervisorBar)
        menubar.add_cascade(label="Supervisor",menu=supervisorMenu)
        employeeMenu = Menu(menubar, tearoff=0)
        employeeMenu.add_command(label = "Employee",command=employeeBar)
        menubar.add_cascade(label="Employee",menu=employeeMenu)
        backupMenu = Menu(menubar, tearoff=0)
        backupMenu.add_command(label = "Backup",command=backupBar)
        menubar.add_cascade(label="Backup",menu=backupMenu)
        inventoryMenu = Menu(menubar, tearoff=0)
        inventoryMenu.add_command(label = "Inventory",command=inventoryBar)
        menubar.add_cascade(label="Inventory",menu=inventoryMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label = "Log Out",command=logOutBar)
        menubar.add_cascade(label="Log Out",menu=logoutMenu)
        master.config(menu=menubar)
        #Basic Home Screen Configuration
        employee = Button(master,text = "Employee",command=employee,width=30,height=5 )
        employee.place(relx=.5, rely=.2, anchor=CENTER)
        departments=Button(master,text = "Departments",command=departments,width=30,height=5)
        departments.place(relx=.5,rely=.5,anchor = CENTER)
        activity=Button(master,text="Activity Log",command=activityLog,width=30,height=5)
        activity.place(relx=.5,rely=.8,anchor=CENTER)
class Supervisor(Frame):
    def __init__(self,master):
        master.title("Admin")
        master.geometry("500x500")
        for widget in master.winfo_children():
            widget.destroy()
        #Admin functions
        def supervisorBar():
            AdminSupervisor(root)
        def backupBar():
            AdminBackup(root)
        def employeeBar():
            AdminEmployee(root)
        def inventoryBar():
            AdminInventory(root)
        def logOutBar():
            Login(root)
        def employee():
            Employee(root)
        def departments():
            Departments(root)
        def activityLog():
            Activity_Log(root)
        #Admin menu bar
        menubar = Menu(master)
        employeeMenu = Menu(menubar, tearoff=0)
        employeeMenu.add_command(label = "Employee",command=employeeBar)
        menubar.add_cascade(label="Employee",menu=employeeMenu)
        inventoryMenu = Menu(menubar, tearoff=0)
        inventoryMenu.add_command(label = "Inventory",command=inventoryBar)
        menubar.add_cascade(label="Inventory",menu=inventoryMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label = "Log Out",command=logOutBar)
        menubar.add_cascade(label="Log Out",menu=logoutMenu)
        master.config(menu=menubar)
        #Basic Home Screen Configuration
        employee = Button(master,text = "Employee",command=employee,width=30,height=5 )
        employee.place(relx=.5, rely=.2, anchor=CENTER)
        departments=Button(master,text = "Departments",command=departments,width=30,height=5)
        departments.place(relx=.5,rely=.5,anchor = CENTER)
        activity=Button(master,text="Activity Log",command=activityLog,width=30,height=5)
        activity.place(relx=.5,rely=.8,anchor=CENTER)
class EmployeeCustomerSatisfaction(Frame):
    def __init__(self,master):
        master.title("Satisfaction Survey")
        master.geometry("500x500")
        for widget in master.winfo_children():
            widget.destroy()
        def Update_Survey():
            global wb
            wb = Workbook()
            wb = load_workbook("C:\\Users\CHR5S\Desktop\Customer Survey.xlsx")

            global ws
            wb.active = 0
            ws = (wb.active)

            if(user_Input1.get() == 'Yes'):
                cell1= ws['B2']
                ws["B2"]= cell1.value+1


            if(user_Input1.get()== 'No'):
                cell1= ws['C2']
                ws["C2"]= cell1.value+1


            if(user_Input2.get()== 'Yes'):
                cell2= ws['B3']
                ws["B3"]= cell2.value+1

            if(user_Input2.get()== 'No'):
                cell2= ws['C3']
                ws["C3"]= cell2.value+1


            if(user_Input3.get()== 'Yes'):
                cell3= ws['B4']
                ws["B4"]= cell3.value+1

            if(user_Input3.get()== 'No'):
                cell3= ws['C4']
                ws["C4"]= cell3.value+1

            wb.save("C:\\Users\CHR5S\Desktop\Customer Survey.xlsx")
                
                # HEY EDWARD LOOK RIGHT HERE
                # DONT INGNORE WHAT YOU CANT HANDLE
                # LOOK AT ME BABY
                # ASK CHRISPOFJET TO NAVIGATE BACK TO EMPLOYEE
                # ALSO THIS CODE IS STOOOPID

        ttk.Label(master, text="Did you find what you were looking for?").grid(column=0, row=5, padx=10, pady=25)
        ttk.Label(master, text="Did you enjoy your visit today?").grid(column=0, row=15, padx=10, pady=25)
        ttk.Label(master, text="Would you recommend us to a friend?").grid(column=0, row=25, padx=10, pady=25)
        global user_Input1
        global user_Input2
        global user_Input3
        user_Choice1= StringVar()
        user_Choice2 = StringVar()
        user_Choice3 = StringVar()
        user_Input1 = ttk.Combobox(master, width=27, textvariable=user_Choice1)
        user_Input1['values'] = ('Yes', 'No')
        user_Input1.grid(column=1, row=5)
        user_Input1.current()

        user_Input2 = ttk.Combobox(master, width=27, textvariable=user_Choice2)
        user_Input2['values'] = ('Yes', 'No')
        user_Input2.grid(column=1, row=15)
        user_Input2.current()

        user_Input3 = ttk.Combobox(master, width=27, textvariable=user_Choice3)
        user_Input3['values'] = ('Yes', 'No')
        user_Input3.grid(column=1, row=25)
        user_Input3.current()


        Survey_Button = Button(master, text="Submit",command=Update_Survey)
        Survey_Button.place(relx=.5,rely=.5,anchor=CENTER)

class Employee(Frame):
    def __init__(self,master):
        master.title("Employee")
        master.geometry("500x500")
        for widget in master.winfo_children():
            widget.destroy()
        #Employee functions
        def logOutBar():
            Login(root)
        def viewInventory():
            Departments(root)
        def activityLog():
            Activity_Log(root)
        def customerSatisfaction():
            EmployeeCustomerSatisfaction(root)
        #Employee menu bar
        menubar = Menu(master)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label = "Log Out",command=logOutBar)
        menubar.add_cascade(label="Log Out",menu=logoutMenu)
        master.config(menu=menubar)
        #Basic Home Screen Configuration
        viewInventory = Button(master,text = "View Inventory",command = viewInventory,width=30,height=5 )
        viewInventory.place(relx=.5, rely=.2, anchor=CENTER)
        satisfactionSurvey=Button(master,text = "Customer Satisfaction Survey",command=customerSatisfaction,width=30,height=5)
        satisfactionSurvey.place(relx=.5,rely=.5,anchor = CENTER)
        employeeActivity=Button(master,text="Activity Log",command=activityLog,width=30,height=5)
        employeeActivity.place(relx=.5,rely=.8,anchor=CENTER)
class Departments(Frame):
    def __init__(self,master):
        master.title("Departments")
        master.geometry("500x500")
        def logOutBar():
            Login(root)
        for widget in master.winfo_children():
            widget.destroy()
        wb = Workbook()
        wb = load_workbook('C:\\Users\CHR5S\Desktop\Hardware.xlsx')

        def set_Sheet_Hardware():
            wb.active = 0
            ws = (wb.active)
            global column_a 
            column_a= ws['A']
            global column_b 
            column_b= ws['B']
            get_info()

        def set_Sheet_Electrical():
            wb.active = 1
            ws = (wb.active)
            global column_a 
            column_a= ws['A']
            global column_b 
            column_b= ws['B']
            get_info()

        def set_Sheet_plumbing():
            wb.active = 2
            ws = (wb.active)
            global column_a 
            column_a= ws['A']
            global column_b 
            column_b= ws['B']
            get_info()
        
        def set_Sheet_Flooring():
            wb.active = 3
            ws = (wb.active)
            global column_a 
            column_a= ws['A']
            global column_b 
            column_b= ws['B']
            get_info()

        def set_Sheet_Lumber():
            wb.active = 4
            ws = (wb.active)
            global column_a 
            column_a= ws['A']
            global column_b 
            column_b= ws['B']
            get_info()
        


        def get_info():
            get_a()
            get_b()
            

            label_a.place(relx=.5,rely=.2,anchor=CENTER)
            label_b.place(relx=.7,rely=.2,anchor=CENTER)
            
        def get_a():
            list =''
            for cell in column_a:
                
                list = f' {list+str(cell.value)}\n'

            label_a.config(text=list) 

        def get_b():
            list =''
            for cell in column_b:
                list = f'{list+str(cell.value)}\n'

            label_b.config(text=list)

        
        Hardware_Button = Button(master, text="View Hardware",command=set_Sheet_Hardware)
        Hardware_Button.place(relx=.15,rely=.1,anchor=CENTER)
        Electrical_Button = Button(master, text="View Electrial",command=set_Sheet_Electrical)
        Electrical_Button.place(relx=.15,rely=.2,anchor=CENTER)
        Plumbing_Button = Button(master, text="View Plumbing",command=set_Sheet_plumbing)
        Plumbing_Button.place(relx=.15,rely=.3,anchor=CENTER)
        Flooring_Button = Button(master, text="View Flooring",command=set_Sheet_Flooring)
        Flooring_Button.place(relx=.15,rely=.4,anchor=CENTER)

        Lumber_Button = Button(master, text="View Lumber",command=set_Sheet_Lumber)
        Lumber_Button.place(relx=.15,rely=.5,anchor=CENTER)

        label_a = Label(master, text="")
        label_b = Label(master, text="")
        #Employee menu bar
        menubar = Menu(master)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label = "Log Out",command=logOutBar)
        menubar.add_cascade(label="Log Out",menu=logoutMenu)
        master.config(menu=menubar)
class Activity_Log():
    def __init__(self,master):
        master.title("Activity Log")
        master.geometry("500x500")
        for widget in master.winfo_children():
            widget.destroy()
         #Update Microsoft Excel Spreadsheet Function
        def updateExcel(x,y,z):
            #Get Size of Activity Log and start count at len(activity_size)
            global count
            df2.loc[count]=[x,y,z]
            with pd.ExcelWriter('C:\\Activity.xlsx') as writer:
                df2.to_excel(writer)
            count+=1
        
        df2=pd.DataFrame(columns=['EmployeeID','Description','Time'])    
        df = pd.read_excel('C:\\Activity.xlsx')
        activityLabel = Label(master,text=df).grid(row=0,column=0)
        employee=StringVar()
        description=StringVar()
        time=StringVar()
        elabel=Label(master,text="Employee ID").grid(row=3,column=0)
        dlabel=Label(master,text="Description").grid(row=4,column=0)
        tlabel=Label(master,text="Time").grid(row=5,column=0)
        e=Entry(master,textvariable=employee,width=20).grid(row=3,column=1)
        d=Entry(master,textvariable=description,width=20).grid(row=4,column=1)
        t=Entry(master,textvariable=time,width=20).grid(row=5,column=1)
        update_button=Button(master,text="Update",command=lambda: updateExcel(employee.get(),description.get(),time.get()))
        update_button.grid(row=6,column=0)
        #Employee menu bar
        menubar = Menu(master)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label = "Log Out",command=logOutBar)
        menubar.add_cascade(label="Log Out",menu=logoutMenu)
        master.config(menu=menubar)

app = Login(root)
root.mainloop()