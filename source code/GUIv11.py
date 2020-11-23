from tkinter import *
import pandas as pd
from numpy import random

import openpyxl as xl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from tkinter import ttk

root = Tk()

admins = {"": ""}
employees = []
supervisors = []

hw_path = "Hardware.xlsx"
activity_path = "Activity.xlsx"
backup_path = "Backup.xlsx"
cs_path = "Customer Survey.xlsx"
dept_path = "Departments.xlsx"

count = 0
labels = []

class Login(Frame):
    def __init__(self, master):
        master.title("User Login")
        master.geometry("500x500")
        for widget in master.winfo_children():
            widget.destroy()

        def validate(x, y):
            if x in admins:
                if admins[x] == y:
                    Admin(root)
                else:
                    invalid = Label(master, text="Invalid Login.", fg='red').place(relx=.5, rely=.8, anchor=CENTER)
            else:
                for obj in employees:
                    if obj.id == x:
                        if obj.password == y:
                            Employee(root, "Employee")
                        else:
                            invalid = Label(master, text="Invalid Login.", fg='red').place(relx=.5, rely=.8,
                                                                                           anchor=CENTER)
                    else:
                        invalid = Label(master, text="Invalid Login.", fg='red').place(relx=.5, rely=.8, anchor=CENTER)
                for obj in supervisors:
                    if obj.id == x:
                        if obj.password == y:
                            Supervisor(root)
                        else:
                            invalid = Label(master, text="Invalid Login.", fg='red').place(relx=.5, rely=.8,
                                                                                           anchor=CENTER)
                    else:
                        invalid = Label(master, text="Invalid Login.", fg='red').place(relx=.5, rely=.8, anchor=CENTER)

        username = StringVar()
        password = StringVar()
        usr_lbl = Label(master, text="ID: ")
        usr_lbl.place(relx=.3, rely=.4, anchor=CENTER)
        usr_text = Entry(master, textvariable=username, width=20)
        usr_text.place(relx=.5, rely=.4, anchor=CENTER)
        pwd_lbl = Label(master, text="Password: ")
        pwd_lbl.place(relx=.3, rely=.5, anchor=CENTER)
        pwd_text = Entry(master, textvariable=password, show="*", width=20)
        pwd_text.place(relx=.5, rely=.5, anchor=CENTER)
        submit = Button(master, text="Login", command=lambda: validate(username.get(), password.get()))
        submit.place(relx=.5, rely=.6, anchor=CENTER)


class AdminSupervisorEdit(Frame):
    def __init__(self, master):
        master.title("Edit Supervisor")
        master.geometry("500x500")
        for widget in master.winfo_children():
            widget.destroy()

        def edit(a, b, c, d, e):
            for obj in supervisors:
                if obj.id == a:
                    if b != '':
                        obj.name = b
                    if c != '':
                        obj.addr = c
                    if d != '':
                        obj.phone = d
                    if e != '':
                        obj.password = e
            Login(root)

        def goHome():
            Admin(root)

        def logOutBar():
            Login(root)

        def delSupervisor(x):
            print('hi')
            for obj in supervisors:
                if obj.id == x:
                    supervisors.remove(obj)



        new_id = StringVar()
        new_name = StringVar()
        new_address = StringVar()
        new_phone = StringVar()
        new_password = StringVar()
        IDLabel = Label(master, text="ID:").grid(column=0, row=0)
        NameLabel = Label(master, text="Name:").grid(column=0, row=1)
        AddrLabel = Label(master, text="Address:").grid(column=0, row=2)
        PhoneLabel = Label(master, text="Phone:").grid(column=0, row=3)
        PwdLabel = Label(master, text="Password:").grid(column=0, row=4)
        IDEntry = Entry(master, textvariable=new_id).grid(column=1, row=0)
        NameEntry = Entry(master, textvariable=new_name).grid(column=1, row=1)
        AddressEntry = Entry(master, textvariable=new_address).grid(column=1, row=2)
        PhoneEntry = Entry(master, textvariable=new_phone).grid(column=1, row=3)
        PasswordEntry = Entry(master, textvariable=new_password).grid(column=1, row=4)
        submit = Button(master, text="Edit Supervisor", command=lambda: edit(new_id.get(), new_name.get(),
                                                                             new_address.get(), new_phone.get(),
                                                                             new_password.get())).grid(column=0, row=5)

        delete = Button(master, text="Delete Supervisor", command=lambda: delSupervisor(new_id.get())).grid(column=1, row=5)
        # Employee menu bar
        menubar = Menu(master)
        goHomeMenu = Menu(menubar, tearoff=0)
        goHomeMenu.add_command(label="Home", command=goHome)
        menubar.add_cascade(label="Home", menu=goHomeMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        master.config(menu=menubar)


class AdminSupervisorView(Frame):
    def __init__(self, master):
        master.title("View Supervisor")
        master.geometry("500x500")

        def logOutBar():
            Login(root)

        def goHome():
            Admin(root)

        for widget in master.winfo_children():
            widget.destroy()
        for obj in supervisors:
            labelx = Label(master, text=obj.id + ", " + obj.name + ", " + obj.addr + ", " + obj.phone).pack()
            # Employee menu bar
        menubar = Menu(master)
        goHomeMenu = Menu(menubar, tearoff=0)
        goHomeMenu.add_command(label="Home", command=goHome)
        menubar.add_cascade(label="Home", menu=goHomeMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        master.config(menu=menubar)


class SupervisorAdd:
    def __init__(self, new_id, new_name, new_addr, new_phone, new_pass):
        self.id = new_id
        self.name = new_name
        self.addr = new_addr
        self.phone = new_phone
        self.password = new_pass


class AdminSupervisorAdd(Frame):
    def __init__(self, master):
        master.title("Add Supervisor")
        master.geometry("500x500")
        x = random.randint(999999)
        for obj in supervisors:
            while obj.id == x:
                x = random.randint(999999)

        def logOutBar():
            Login(root)

        def goHome():
            Admin(root)

        for widget in master.winfo_children():
            widget.destroy()

        def creates(a, b, c, d, e):
            supervisors.append(EmployeeAdd(a, b, c, d, e))
            Login(root)

        password2 = StringVar()
        name = StringVar()
        address = StringVar()
        phoneNumber = StringVar()
        id_lbl = Label(master, text="ID: " + str(x))
        id_lbl.place(relx=.3, rely=.3, anchor=CENTER)

        pwd_lbl = Label(master, text="Password: ")
        pwd_lbl.place(relx=.3, rely=.4, anchor=CENTER)
        pwd_text = Entry(master, textvariable=password2, show="*", width=20)
        pwd_text.place(relx=.5, rely=.4, anchor=CENTER)

        name_lbl = Label(master, text="Name: ")
        name_lbl.place(relx=.3, rely=.5, anchor=CENTER)
        name_text = Entry(master, textvariable=name, width=20)
        name_text.place(relx=.5, rely=.5, anchor=CENTER)

        addr_lbl = Label(master, text="Address: ")
        addr_lbl.place(relx=.3, rely=.6, anchor=CENTER)
        addr_text = Entry(master, textvariable=address, width=20)
        addr_text.place(relx=.5, rely=.6, anchor=CENTER)

        phone_lbl = Label(master, text="Phone Number: ")
        phone_lbl.place(relx=.3, rely=.7, anchor=CENTER)
        phone_text = Entry(master, textvariable=phoneNumber, width=20)
        phone_text.place(relx=.5, rely=.7, anchor=CENTER)

        createAccount = Button(master, text="Create Account",
                               command=lambda: creates(str(x), name.get(), address.get(),
                                                       phoneNumber.get(), password2.get()))
        createAccount.place(relx=.3, rely=.8, anchor=CENTER)
        # Employee menu bar
        menubar = Menu(master)
        goHomeMenu = Menu(menubar, tearoff=0)
        goHomeMenu.add_command(label="Home", command=goHome)
        menubar.add_cascade(label="Home", menu=goHomeMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        master.config(menu=menubar)


# Substitute Supervisor Edit for Supervisor delete like we have in employee
class AdminSupervisor(Frame):
    def __init__(self, master):
        master.title("Edit Supervisor")
        master.geometry("500x500")

        def logOutBar():
            Login(root)

        def goHome():
            Admin(root)

        def adminSupervisorView():
            AdminSupervisorView(root)

        def adminSupervisorAdd():
            AdminSupervisorAdd(root)

        def adminSupervisorEdit():
            AdminSupervisorEdit(root)

        for widget in master.winfo_children():
            widget.destroy()
        view = Button(master, text="View", command=adminSupervisorView, width=30, height=5)
        view.place(relx=.5, rely=.2, anchor=CENTER)
        add = Button(master, text="Add", command=adminSupervisorAdd, width=30, height=5)
        add.place(relx=.5, rely=.5, anchor=CENTER)
        delete = Button(master, text="Edit/Delete", command=adminSupervisorEdit, width=30, height=5)
        delete.place(relx=.5, rely=.8, anchor=CENTER)
        # Employee menu bar
        menubar = Menu(master)
        goHomeMenu = Menu(menubar, tearoff=0)
        goHomeMenu.add_command(label="Home", command=goHome)
        menubar.add_cascade(label="Home", menu=goHomeMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        master.config(menu=menubar)


class AdminEmployeeDelete(Frame):
    def __init__(self, master):
        master.title("Delete Employee")
        master.geometry("500x500")

        def logOutBar():
            Login(root)

        def goHome():
            Admin(root)

        for widget in master.winfo_children():
            widget.destroy()

        def deletes(x):
            for obj in employees:
                if obj.id == x:
                    employees.remove(obj)
            Login(root)

        username2 = StringVar()
        usr_lbl = Label(master, text="ID: ")
        usr_lbl.place(relx=.3, rely=.4, anchor=CENTER)
        usr_text = Entry(master, textvariable=username2, width=20)
        usr_text.place(relx=.5, rely=.4, anchor=CENTER)
        deleteSupervisor = Button(master, text="Delete Employee", command=lambda: deletes(username2.get()))
        deleteSupervisor.place(relx=.3, rely=.6, anchor=CENTER)

        # Employee menu bar
        menubar = Menu(master)
        goHomeMenu = Menu(menubar, tearoff=0)
        goHomeMenu.add_command(label="Home", command=goHome)
        menubar.add_cascade(label="Home", menu=goHomeMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        master.config(menu=menubar)


class AdminEmployeeView(Frame):
    def __init__(self, master):
        master.title("View Employee")
        master.geometry("500x500")

        def logOutBar():
            Login(root)

        def goHome():
            Admin(root)

        for widget in master.winfo_children():
            widget.destroy()
        for obj in employees:
            labelx = Label(master, text=obj.id + ", " + obj.name + ", " + obj.addr + ", " + obj.phone).pack()
            # Employee menu bar
        menubar = Menu(master)
        goHomeMenu = Menu(menubar, tearoff=0)
        goHomeMenu.add_command(label="Home", command=goHome)
        menubar.add_cascade(label="Home", menu=goHomeMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        master.config(menu=menubar)


class EmployeeAdd:
    def __init__(self, new_id, new_name, new_addr, new_phone, new_pass):
        self.id = new_id
        self.name = new_name
        self.addr = new_addr
        self.phone = new_phone
        self.password = new_pass


class AdminEmployeeAdd(Frame):
    def __init__(self, master):
        master.title("Add Employee")
        master.geometry("500x500")
        x = random.randint(999999)
        for obj in employees:
            while obj.id == x:
                x = random.randint(999999)

        def logOutBar():
            Login(root)

        def goHome():
            Admin(root)

        for widget in master.winfo_children():
            widget.destroy()

        def creates(a, b, c, d, e):
            employees.append(EmployeeAdd(a, b, c, d, e))
            Login(root)

        password2 = StringVar()
        name = StringVar()
        address = StringVar()
        phoneNumber = StringVar()
        id_lbl = Label(master, text="ID: " + str(x))
        id_lbl.place(relx=.3, rely=.3, anchor=CENTER)

        pwd_lbl = Label(master, text="Password: ")
        pwd_lbl.place(relx=.3, rely=.4, anchor=CENTER)
        pwd_text = Entry(master, textvariable=password2, show="*", width=20)
        pwd_text.place(relx=.5, rely=.4, anchor=CENTER)

        name_lbl = Label(master, text="Name: ")
        name_lbl.place(relx=.3, rely=.5, anchor=CENTER)
        name_text = Entry(master, textvariable=name, width=20)
        name_text.place(relx=.5, rely=.5, anchor=CENTER)

        addr_lbl = Label(master, text="Address: ")
        addr_lbl.place(relx=.3, rely=.6, anchor=CENTER)
        addr_text = Entry(master, textvariable=address, width=20)
        addr_text.place(relx=.5, rely=.6, anchor=CENTER)

        phone_lbl = Label(master, text="Phone Number: ")
        phone_lbl.place(relx=.3, rely=.7, anchor=CENTER)
        phone_text = Entry(master, textvariable=phoneNumber, width=20)
        phone_text.place(relx=.5, rely=.7, anchor=CENTER)

        createAccount = Button(master, text="Create Account",
                               command=lambda: creates(str(x), name.get(), address.get(),
                                                       phoneNumber.get(), password2.get()))
        createAccount.place(relx=.3, rely=.8, anchor=CENTER)
        # Employee menu bar
        menubar = Menu(master)
        goHomeMenu = Menu(menubar, tearoff=0)
        goHomeMenu.add_command(label="Home", command=goHome)
        menubar.add_cascade(label="Home", menu=goHomeMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        master.config(menu=menubar)


# Does not change phone number
class AdminEmployeeEdit(Frame):
    def __init__(self, master):
        master.title("Edit Employee")
        master.geometry("500x500")
        for widget in master.winfo_children():
            widget.destroy()

        def edit(a, b, c, d, e):
            for obj in employees:
                if obj.id == a:
                    if b != '':
                        obj.name = b
                    if c != '':
                        obj.addr = c
                    if d != '':
                        obj.phone = d
                    if e != '':
                        obj.password = e
            Login(root)

        def goHome():
            Admin(root)

        def logOutBar():
            Login(root)

        def delEmployee(x):
            for obj in employees:
                if obj.id == x:
                    employees.remove(obj)

        new_id = StringVar()
        new_name = StringVar()
        new_address = StringVar()
        new_phone = StringVar()
        new_password = StringVar()
        IDLabel = Label(master, text="ID:").grid(column=0, row=0)
        NameLabel = Label(master, text="Name:").grid(column=0, row=1)
        AddrLabel = Label(master, text="Address:").grid(column=0, row=2)
        PhoneLabel = Label(master, text="Phone:").grid(column=0, row=3)
        PwdLabel = Label(master, text="Password:").grid(column=0, row=4)
        IDEntry = Entry(master, textvariable=new_id).grid(column=1, row=0)
        NameEntry = Entry(master, textvariable=new_name).grid(column=1, row=1)
        AddressEntry = Entry(master, textvariable=new_address).grid(column=1, row=2)
        PhoneEntry = Entry(master, textvariable=new_phone).grid(column=1, row=3)
        PasswordEntry = Entry(master, textvariable=new_password).grid(column=1, row=4)
        submit = Button(master, text="Edit Employee",
                        command=lambda: edit(new_id.get(), new_name.get(), new_address.get(), new_phone.get(),
                                             new_password.get())).grid(column=0, row=5)
        delete = Button(master, text="Delete Employee",command=lambda: delEmployee(new_id.get())).grid(row=5, column=1)
        # Employee menu bar
        menubar = Menu(master)
        goHomeMenu = Menu(menubar, tearoff=0)
        goHomeMenu.add_command(label="Home", command=goHome)
        menubar.add_cascade(label="Home", menu=goHomeMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        master.config(menu=menubar)


class AdminEmployee(Frame):
    def __init__(self, master):
        master.title("Edit Employee")
        master.geometry("500x500")

        def logOutBar():
            Login(root)

        def goHome():
            Admin(root)

        def adminEmployeeView():
            AdminEmployeeView(root)

        def adminEmployeeAdd():
            AdminEmployeeAdd(root)

        def adminEmployeeDelete():
            AdminEmployeeDelete(root)

        def adminEmployeeEdit():
            AdminEmployeeEdit(root)

        for widget in master.winfo_children():
            widget.destroy()
        view = Button(master, text="View", command=adminEmployeeView, width=30, height=5)
        view.place(relx=.5, rely=.2, anchor=CENTER)
        add = Button(master, text="Add", command=adminEmployeeAdd, width=30, height=5)
        add.place(relx=.5, rely=.5, anchor=CENTER)
        # Delete could possible be just a scenario of edit
        # May possibly delete adminEmployeeDelete class
        # delete=Button(master,text="Delete",command=adminEmployeeDelete,width=30,height=5)
        # delete.place(relx=.5,rely=.8,anchor=CENTER)
        edit = Button(master, text="Edit/Delete", command=adminEmployeeEdit, width=30, height=5)
        edit.place(relx=.5, rely=.8, anchor=CENTER)
        # Employee menu bar
        menubar = Menu(master)
        goHomeMenu = Menu(menubar, tearoff=0)
        goHomeMenu.add_command(label="Home", command=goHome)
        menubar.add_cascade(label="Home", menu=goHomeMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        master.config(menu=menubar)


class AdminBackup(Frame):
    def __init__(self, master):
        master.title("Edit Backup")
        master.geometry("500x500")

        def logOutBar():
            Login(root)

        def goHome():
            Admin(root)

        for widget in master.winfo_children():
            widget.destroy()

        def createBackup():
            wb1 = xl.load_workbook(hw_path)
            wb2=xl.load_workbook(backup_path)
            def writeToExcel(wb1,ws1,wb2,ws2,mr,mc):
                # copying the cell values from source  
                # excel file to destination excel file 
                for i in range (1, mr + 1): 
                    for j in range (1, mc + 1): 
                        # reading cell value from source excel file 
                        c = ws1.cell(row = i, column = j) 
  
                        # writing the read value to destination excel file 
                        ws2.cell(row = i, column = j).value = c.value 
  
                # saving the destination excel file 
                wb2.save(str(backup_path)) 
                
            for index in range(0,5):
                ws1 = wb1.worksheets[index]
                ws2 = wb2.worksheets[index]
                mr = ws1.max_row
                mc = ws1.max_column
                writeToExcel(wb1,ws1,wb2,ws2,mr,mc)
            completed = Label(master, text = "Backup created").pack()
            
                
                
                    
        def restoreBackup():
            wb1 = xl.load_workbook(backup_path)
            wb2 = xl.load_workbook(hw_path)
            
            def writeToExcel(wb1,ws1,wb2,ws2,mr,mc):
                # copying the cell values from source  
                # excel file to destination excel file 
                for i in range (1, mr + 1): 
                    for j in range (1, mc + 1): 
                        # reading cell value from source excel file 
                        c = ws1.cell(row = i, column = j) 
  
                        # writing the read value to destination excel file 
                        ws2.cell(row = i, column = j).value = c.value 
  
                # saving the destination excel file 
                wb2.save(str(hw_path)) 
            
            for index in range(0,5):
                ws1 = wb1.worksheets[index]
                ws2 = wb2.worksheets[index]
                mr = ws1.max_row
                mc = ws1.max_column
                writeToExcel(wb1,ws1,wb2,ws2,mr,mc)
            complete = Label(master, text = "Restored from backup").pack()
            
            

        view = Button(master, text="Create", command=createBackup, width=30, height=5)
        view.place(relx=.5, rely=.2, anchor=CENTER)
        add = Button(master, text="Restore", command=restoreBackup, width=30, height=5)
        add.place(relx=.5, rely=.5, anchor=CENTER)
        # Employee menu bar
        menubar = Menu(master)
        goHomeMenu = Menu(menubar, tearoff=0)
        goHomeMenu.add_command(label="Home", command=goHome)
        menubar.add_cascade(label="Home", menu=goHomeMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        master.config(menu=menubar)

#####
#####
#####
#####
class AdminInventoryUpdate(Frame):
    def __init__(self, master):
        master.title("Update/Delete Inventory")
        master.geometry("500x500")
        for widget in master.winfo_children():
            widget.destroy()
        wb = Workbook()
        wb = load_workbook(hw_path)
        labels = []
        def up_Hardware():
            wb.active = 0
            ws = (wb.active)
            global column_a
            global column_b
            column_a = ws['A']
            column_b = ws['B']
            get_info(wb.active)

        def up_Electrical():
            wb.active = 1
            ws = (wb.active)
            global column_a
            global column_b
            column_a = ws['A']
            column_b = ws['B']
            get_info(wb.active)

        def up_Plumbing():
            wb.active = 2
            ws = (wb.active)
            global column_a
            global column_b
            column_a = ws['A']
            column_b = ws['B']
            get_info(wb.active)

        def up_Flooring():
            wb.active = 3
            ws = (wb.active)
            global column_a
            global column_b
            column_a = ws['A']
            column_b = ws['B']
            get_info(wb.active)

        def up_Lumber():
            wb.active = 4
            ws = (wb.active)
            global column_a
            global column_b
            column_a = ws['A']
            column_b = ws['B']
            get_info(wb.active)
        
        def printCell(x,y):
            #x is a list of entries, y is the active workbook
            for col in y.iter_cols(min_col=2,max_col=2):
                for i,cell in enumerate(col):
                    #print(str(i)+ ": " + str(cell.value))
                    for index,entry in enumerate(x):
                        if(i==index):
                            cell.value=entry.get()
            wb.save(hw_path)
        def get_info(x):
            for label in labels:
                label.destroy()
            #print("Active workbook is " + str(x))
            list = ''
            entries = []
            for index, cell in enumerate(column_a,start=0):
                list = f' {list + str(cell.value)}\n'
                newInventory = Label(master,text=cell.value)
                newInventory.grid(row=index,column=2)
                labels.append(newInventory)
                newInventory2 = Entry(master)
                newInventory2.grid(row=index,column=3)
                newInventory2.insert(0,column_b[index].value)
                entries.append(newInventory2)
            btn = Button(master, text = "Update",command=lambda:printCell(entries,x)).grid(row=10,column=2)

        Hardware_Button = Button(master, text="Update Hardware", command=up_Hardware).grid(row=0,column=0)
        Electrical_Button = Button(master, text="Update Electrical", command=up_Electrical).grid(row=1,column=0)
        Plumbing_Button = Button(master, text="Update Plumbing", command=up_Plumbing).grid(row=2,column=0)
        Flooring_Button = Button(master, text="Update Flooring", command=up_Flooring).grid(row=3,column=0)
        Lumber_Button = Button(master, text="Update Lumber", command=up_Lumber).grid(row=4,column=0)


class AdminInventory(Frame):
    def __init__(self, master):
        master.title("Edit Inventory")
        master.geometry("500x500")

        def logOutBar():
            Login(root)

        def viewInventory():
            Departments(root, "Admin")

        def goHome():
            Admin(root)

        def updateInventory():
            AdminInventoryUpdate(root)

        for widget in master.winfo_children():
            widget.destroy()
        # Adding functions for Update and Delete Inventory
        #
        view = Button(master, text="View", command=viewInventory, width=30, height=5)
        view.place(relx=.5, rely=.2, anchor=CENTER)
        add = Button(master, text="Update", command=updateInventory, width=30, height=5)
        add.place(relx=.5, rely=.5, anchor=CENTER)

        # Employee menu bar
        menubar = Menu(master)
        goHomeMenu = Menu(menubar, tearoff=0)
        goHomeMenu.add_command(label="Home", command=goHome)
        menubar.add_cascade(label="Home", menu=goHomeMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        master.config(menu=menubar)


class AdminDepartmentCreate(Frame):
    def __init__(self, master):
        master.title("Edit Departments")
        master.geometry("500x500")
        for widget in master.winfo_children():
            widget.destroy()
        departmentName = StringVar()

        def goHome():
            Admin(root)

        def logOutBar():
            Login(root)

        def creates(x):
            df = pd.read_excel(dept_path)
            num = len(df)
            df.loc[num, 'Name'] = x
            df.loc[num, 'ID'] = num + 1
            df.loc[num, 'Visible'] = True
            with pd.ExcelWriter(dept_path) as writer:
                df.to_excel(writer)
            Admin(root)

        l = Label(master, text="Department name:")
        l.place(relx=.2, rely=.3)
        e = Entry(master, textvariable=departmentName)
        e.place(relx=.5, rely=.3)
        create = Button(master, text="Create Department", command=lambda: creates(departmentName.get()))
        create.place(relx=.3, rely=.5)
        # Employee menu bar
        menubar = Menu(master)
        goHomeMenu = Menu(menubar, tearoff=0)
        goHomeMenu.add_command(label="Home", command=goHome)
        menubar.add_cascade(label="Home", menu=goHomeMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        master.config(menu=menubar)


class AdminDepartmentDelete(Frame):
    def __init__(self, master):
        master.title("Delete Departments")
        master.geometry("500x500")
        for widget in master.winfo_children():
            widget.destroy()

        def goHome():
            Admin(root)

        def logOutBar():
            Login(root)

        # Write an algorithm that sets the Visible value of the selected Department to False
        def deletes(x):
            df = pd.read_excel(dept_path)
            # look for row containing bathroom
            for index, row in df.iterrows():
                if row['Name'] == x:
                    df.loc[index, 'Visible'] = False

            with pd.ExcelWriter(dept_path) as writer:
                df.to_excel(writer)
            Admin(root)

        departmentName = StringVar()
        l = Label(master, text="Department name:")
        l.place(relx=.2, rely=.3)
        e = Entry(master, textvariable=departmentName)
        e.place(relx=.5, rely=.3)
        create = Button(master, text="Delete Department", command=lambda: deletes(departmentName.get()))
        create.place(relx=.3, rely=.5)

        # Employee menu bar
        menubar = Menu(master)
        goHomeMenu = Menu(menubar, tearoff=0)
        goHomeMenu.add_command(label="Home", command=goHome)
        menubar.add_cascade(label="Home", menu=goHomeMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        master.config(menu=menubar)


class AdminDepartment(Frame):
    def __init__(self, master):
        master.title("Edit Departments")
        master.geometry("500x500")
        for widget in master.winfo_children():
            widget.destroy()

        def goHome():
            Admin(root)

        def logOutBar():
            Login(root)

        def createDepartment():
            AdminDepartmentCreate(root)

        def deleteDepartment():
            AdminDepartmentDelete(root)

        # Basic Home Screen Configuration
        employee = Button(master, text="Create", command=createDepartment, width=30, height=5)
        employee.place(relx=.5, rely=.2, anchor=CENTER)
        departments = Button(master, text="Delete", command=deleteDepartment, width=30, height=5)
        departments.place(relx=.5, rely=.5, anchor=CENTER)
        # Employee menu bar
        menubar = Menu(master)
        goHomeMenu = Menu(menubar, tearoff=0)
        goHomeMenu.add_command(label="Home", command=goHome)
        menubar.add_cascade(label="Home", menu=goHomeMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        master.config(menu=menubar)


# Where do we keep our transactions?
# Transactions is its own spreadsheet
# Transactions has a transactionID, dollar amount, and date
class AdminTransactions(Frame):
    def __init__(self, master):
        master.title("Transactions")
        master.geometry("500x500")
        for widget in master.winfo_children():
            widget.destroy()
        employee = Button(master, text="View", width=30, height=5)
        employee.place(relx=.5, rely=.2, anchor=CENTER)
        departments = Button(master, text="Delete", width=30, height=5)
        departments.place(relx=.5, rely=.5, anchor=CENTER)


class Admin(Frame):
    def __init__(self, master):
        master.title("Admin")
        master.geometry("500x500")
        for widget in master.winfo_children():
            widget.destroy()

        # Admin functions
        def supervisorBar():
            AdminSupervisor(root)

        def backupBar():
            AdminBackup(root)

        def transactionsBar():
            AdminTransactions(root)

        def employeeBar():
            AdminEmployee(root)

        def inventoryBar():
            AdminInventory(root)

        def departmentBar():
            AdminDepartment(root)

        def logOutBar():
            Login(root)

        def employee():
            Employee(root, "Admin")

        def departments():
            Departments(root, "Admin")

        def activityLog():
            Activity_Log(root, "Admin")

        # Admin menu bar
        menubar = Menu(master)
        supervisorMenu = Menu(menubar, tearoff=0)
        supervisorMenu.add_command(label="Supervisor", command=supervisorBar)
        menubar.add_cascade(label="Supervisor", menu=supervisorMenu)
        employeeMenu = Menu(menubar, tearoff=0)
        employeeMenu.add_command(label="Employee", command=employeeBar)
        menubar.add_cascade(label="Employee", menu=employeeMenu)
        backupMenu = Menu(menubar, tearoff=0)
        backupMenu.add_command(label="Backup", command=backupBar)
        menubar.add_cascade(label="Backup", menu=backupMenu)
        inventoryMenu = Menu(menubar, tearoff=0)
        inventoryMenu.add_command(label="Inventory", command=inventoryBar)
        menubar.add_cascade(label="Inventory", menu=inventoryMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        departmentMenu = Menu(menubar, tearoff=0)
        departmentMenu.add_command(label="Department", command=departmentBar)
        menubar.add_cascade(label="Department", menu=departmentMenu)
        transactionsMenu = Menu(menubar, tearoff=0)
        transactionsMenu.add_command(label="Transactions", command=transactionsBar)
        menubar.add_cascade(label="Transactions", menu=transactionsMenu)
        master.config(menu=menubar)
        # Basic Home Screen Configuration
        employee = Button(master, text="Employee", command=employee, width=30, height=5)
        employee.place(relx=.5, rely=.2, anchor=CENTER)
        departments = Button(master, text="Departments", command=departments, width=30, height=5)
        departments.place(relx=.5, rely=.5, anchor=CENTER)
        activity = Button(master, text="Activity Log", command=activityLog, width=30, height=5)
        activity.place(relx=.5, rely=.8, anchor=CENTER)


class SupervisorEmployeeDelete(Frame):
    def __init__(self, master):
        master.title("Delete Employee")
        master.geometry("500x500")

        def logOutBar():
            Login(root)

        def goHome():
            Supervisor(root)

        for widget in master.winfo_children():
            widget.destroy()

        def deletes(x):
            for obj in employees:
                if obj.id == x:
                    employees.remove(obj)
            Login(root)

        username2 = StringVar()
        usr_lbl = Label(master, text="ID: ")
        usr_lbl.place(relx=.3, rely=.4, anchor=CENTER)
        usr_text = Entry(master, textvariable=username2, width=20)
        usr_text.place(relx=.5, rely=.4, anchor=CENTER)
        deleteSupervisor = Button(master, text="Delete Employee", command=lambda: deletes(username2.get()))
        deleteSupervisor.place(relx=.3, rely=.6, anchor=CENTER)

        # Employee menu bar
        menubar = Menu(master)
        goHomeMenu = Menu(menubar, tearoff=0)
        goHomeMenu.add_command(label="Home", command=goHome)
        menubar.add_cascade(label="Home", menu=goHomeMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        master.config(menu=menubar)


class SupervisorEmployeeAdd(Frame):
    def __init__(self, master):
        master.title("Add Employee")
        master.geometry("500x500")
        x = random.randint(999999)
        for obj in employees:
            while obj.id == x:
                x = random.randint(999999)

        def logOutBar():
            Login(root)

        def goHome():
            Supervisor(root)

        for widget in master.winfo_children():
            widget.destroy()

        def creates(a, b, c, d, e):
            employees.append(EmployeeAdd(a, b, c, d, e))
            Login(root)

        password2 = StringVar()
        name = StringVar()
        address = StringVar()
        phoneNumber = StringVar()
        id_lbl = Label(master, text="ID: " + str(x))
        id_lbl.place(relx=.3, rely=.3, anchor=CENTER)

        pwd_lbl = Label(master, text="Password: ")
        pwd_lbl.place(relx=.3, rely=.4, anchor=CENTER)
        pwd_text = Entry(master, textvariable=password2, show="*", width=20)
        pwd_text.place(relx=.5, rely=.4, anchor=CENTER)

        name_lbl = Label(master, text="Name: ")
        name_lbl.place(relx=.3, rely=.5, anchor=CENTER)
        name_text = Entry(master, textvariable=name, width=20)
        name_text.place(relx=.5, rely=.5, anchor=CENTER)

        addr_lbl = Label(master, text="Address: ")
        addr_lbl.place(relx=.3, rely=.6, anchor=CENTER)
        addr_text = Entry(master, textvariable=address, width=20)
        addr_text.place(relx=.5, rely=.6, anchor=CENTER)

        phone_lbl = Label(master, text="Phone Number: ")
        phone_lbl.place(relx=.3, rely=.7, anchor=CENTER)
        phone_text = Entry(master, textvariable=phoneNumber, width=20)
        phone_text.place(relx=.5, rely=.7, anchor=CENTER)

        createAccount = Button(master, text="Create Account",
                               command=lambda: creates(str(x), name.get(), address.get(),
                                                       phoneNumber.get(), password2.get()))
        createAccount.place(relx=.3, rely=.8, anchor=CENTER)
        # Employee menu bar
        menubar = Menu(master)
        goHomeMenu = Menu(menubar, tearoff=0)
        goHomeMenu.add_command(label="Home", command=goHome)
        menubar.add_cascade(label="Home", menu=goHomeMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        master.config(menu=menubar)


class SupervisorEmployeeView(Frame):
    def __init__(self, master):
        master.title("View Employee")
        master.geometry("500x500")

        def logOutBar():
            Login(root)

        def goHome():
            Supervisor(root)

        for widget in master.winfo_children():
            widget.destroy()
        for obj in employees:
            labelx = Label(master, text=obj.id).pack()
            # Employee menu bar
        menubar = Menu(master)
        goHomeMenu = Menu(menubar, tearoff=0)
        goHomeMenu.add_command(label="Home", command=goHome)
        menubar.add_cascade(label="Home", menu=goHomeMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        master.config(menu=menubar)


class SupervisorEmployee(Frame):
    def __init__(self, master):
        master.title("Edit Employee")
        master.geometry("500x500")

        def logOutBar():
            Login(root)

        def goHome():
            Supervisor(root)

        def supervisorEmployeeView():
            SupervisorEmployeeView(root)

        def supervisorEmployeeAdd():
            SupervisorEmployeeAdd(root)

        def supervisorEmployeeDelete():
            SupervisorEmployeeDelete(root)

        for widget in master.winfo_children():
            widget.destroy()
        view = Button(master, text="View", command=supervisorEmployeeView, width=30, height=5)
        view.place(relx=.5, rely=.2, anchor=CENTER)
        add = Button(master, text="Add", command=supervisorEmployeeAdd, width=30, height=5)
        add.place(relx=.5, rely=.5, anchor=CENTER)
        delete = Button(master, text="Delete", command=supervisorEmployeeDelete, width=30, height=5)
        delete.place(relx=.5, rely=.8, anchor=CENTER)
        # Employee menu bar
        menubar = Menu(master)
        goHomeMenu = Menu(menubar, tearoff=0)
        goHomeMenu.add_command(label="Home", command=goHome)
        menubar.add_cascade(label="Home", menu=goHomeMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        master.config(menu=menubar)


class Supervisor(Frame):
    def __init__(self, master):
        master.title("Supervisor")
        master.geometry("500x500")
        for widget in master.winfo_children():
            widget.destroy()

        # Supervisor functions
        def employeeBar():
            SupervisorEmployee(root)

        def inventoryBar():
            AdminInventory(root)

        def logOutBar():
            Login(root)

        def employee():
            Employee(root, "Supervisor")

        def departments():
            Departments(root, "Supervisor")

        def activityLog():
            Activity_Log(root, "Supervisor")

        # Admin menu bar
        menubar = Menu(master)
        employeeMenu = Menu(menubar, tearoff=0)
        employeeMenu.add_command(label="Employee", command=employeeBar)
        menubar.add_cascade(label="Employee", menu=employeeMenu)
        inventoryMenu = Menu(menubar, tearoff=0)
        inventoryMenu.add_command(label="Inventory", command=inventoryBar)
        menubar.add_cascade(label="Inventory", menu=inventoryMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        master.config(menu=menubar)
        # Basic Home Screen Configuration
        employee = Button(master, text="Employee", command=employee, width=30, height=5)
        employee.place(relx=.5, rely=.2, anchor=CENTER)
        departments = Button(master, text="Departments", command=departments, width=30, height=5)
        departments.place(relx=.5, rely=.5, anchor=CENTER)
        activity = Button(master, text="Activity Log", command=activityLog, width=30, height=5)
        activity.place(relx=.5, rely=.8, anchor=CENTER)


class EmployeeCustomerSatisfaction(Frame):
    def __init__(self, master):
        master.title("Satisfaction Survey")
        master.geometry("500x500")
        for widget in master.winfo_children():
            widget.destroy()

        # Delete customer satisfaction rating
        def Cancel_Survey():
            Employee(root, "Employee")

        def logOutBar():
            Login(root)

        def surveyRest():
            wb = Workbook()
            wb = load_workbook(cs_path)

            wb.active = 0
            ws = (wb.active)

            ws['B2'] = 0
            ws['B3'] = 0
            ws['B4'] = 0
            ws['C2'] = 0
            ws['C3'] = 0
            ws['C4'] = 0
            wb.save(cs_path)

        def Update_Survey():
            global wb
            wb = Workbook()
            wb = load_workbook(cs_path)

            global ws
            wb.active = 0
            ws = (wb.active)

            if (user_Input1.get() == 'Yes'):
                cell1 = ws['B2']
                ws["B2"] = cell1.value + 1

            if (user_Input1.get() == 'No'):
                cell1 = ws['C2']
                ws["C2"] = cell1.value + 1

            if (user_Input2.get() == 'Yes'):
                cell2 = ws['B3']
                ws["B3"] = cell2.value + 1

            if (user_Input2.get() == 'No'):
                cell2 = ws['C3']
                ws["C3"] = cell2.value + 1

            if (user_Input3.get() == 'Yes'):
                cell3 = ws['B4']
                ws["B4"] = cell3.value + 1

            if (user_Input3.get() == 'No'):
                cell3 = ws['C4']
                ws["C4"] = cell3.value + 1

            # Customer Satisfaction Survey should display score
            print(cell1.value + cell2.value + cell3.value)

            wb.save(cs_path)

        ttk.Label(master, text="Did you find what you were looking for?").grid(column=0, row=5, padx=10, pady=25)
        ttk.Label(master, text="Did you enjoy your visit today?").grid(column=0, row=15, padx=10, pady=25)
        ttk.Label(master, text="Would you recommend us to a friend?").grid(column=0, row=25, padx=10, pady=25)
        global user_Input1
        global user_Input2
        global user_Input3
        user_Choice1 = StringVar()
        user_Choice2 = StringVar()
        user_Choice3 = StringVar()
        user_Input1 = ttk.Combobox(master, width=27, textvariable=user_Choice1)
        user_Input1['values'] = ('Yes', 'No')
        user_Input1.grid(column=1, row=5)
        user_Input1.current()

        user_Input2 = ttk.Combobox(master, width=27, textvariable=user_Choice2)
        user_Input2['values'] = ('Yes', 'No')
        user_Input2.grid(column=1, row=15)
        user_Input2.current()

        user_Input3 = ttk.Combobox(master, width=27, textvariable=user_Choice3)
        user_Input3['values'] = ('Yes', 'No')
        user_Input3.grid(column=1, row=25)
        user_Input3.current()

        Survey_Button = Button(master, text="Submit", command=Update_Survey)
        Survey_Button.place(relx=.5, rely=.5, anchor=CENTER)
        Cancel_Button = Button(master, text="Reset Survey", command=surveyRest)
        Cancel_Button.place(relx=.7, rely=.5, anchor=CENTER)

        # Customer Satisfaction Survey needs ID of current employee

        # Employee menu bar
        menubar = Menu(master)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        master.config(menu=menubar)


class Employee(Frame):
    def __init__(self, master, who):
        master.title("Employee")
        master.geometry("500x500")
        for widget in master.winfo_children():
            widget.destroy()

        # Employee functions
        def logOutBar():
            Login(root)

        def goHome():
            if who == "Admin":
                Admin(root)
            elif who == "Supervisor":
                Supervisor(root)
            else:
                Employee(root, "Employee")

        def viewInventory():
            Departments(root, "Employee")

        def activityLog():
            Activity_Log(root, "Employee")

        def customerSatisfaction():
            EmployeeCustomerSatisfaction(root)

        # Employee menu bar
        menubar = Menu(master)
        goHomeMenu = Menu(menubar, tearoff=0)
        goHomeMenu.add_command(label="Home", command=goHome)
        menubar.add_cascade(label="Home", menu=goHomeMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        master.config(menu=menubar)
        # Basic Home Screen Configuration
        viewInventory = Button(master, text="View Inventory", command=viewInventory, width=30, height=5)
        viewInventory.place(relx=.5, rely=.2, anchor=CENTER)
        satisfactionSurvey = Button(master, text="Customer Satisfaction Survey", command=customerSatisfaction, width=30,
                                    height=5)
        satisfactionSurvey.place(relx=.5, rely=.5, anchor=CENTER)
        employeeActivity = Button(master, text="Activity Log", command=activityLog, width=30, height=5)
        employeeActivity.place(relx=.5, rely=.8, anchor=CENTER)


class Departments(Frame):
    def __init__(self, master, who):
        master.title("Departments")
        master.geometry("500x500")

        def logOutBar():
            Login(root)

        def goHome():
            if who == "Admin":
                Admin(root)
            elif who == "Supervisor":
                Supervisor(root)
            else:
                Employee(root, "Employee")

        for widget in master.winfo_children():
            widget.destroy()
        wb = Workbook()
        wb = load_workbook(hw_path)

        def set_Sheet_Hardware():
            wb.active = 0
            ws = (wb.active)
            global column_a
            column_a = ws['A']
            global column_b
            column_b = ws['B']
            get_info()

        def set_Sheet_Electrical():
            wb.active = 1
            ws = (wb.active)
            global column_a
            column_a = ws['A']
            global column_b
            column_b = ws['B']
            get_info()

        def set_Sheet_plumbing():
            wb.active = 2
            ws = (wb.active)
            global column_a
            column_a = ws['A']
            global column_b
            column_b = ws['B']
            get_info()

        def set_Sheet_Flooring():
            wb.active = 3
            ws = (wb.active)
            global column_a
            column_a = ws['A']
            global column_b
            column_b = ws['B']
            get_info()

        def set_Sheet_Lumber():
            wb.active = 4
            ws = (wb.active)
            global column_a
            column_a = ws['A']
            global column_b
            column_b = ws['B']
            get_info()

        def get_info():
            get_a()
            get_b()

            label_a.place(relx=.5, rely=.2, anchor=CENTER)
            label_b.place(relx=.7, rely=.2, anchor=CENTER)

        def get_a():
            list = ''
            for cell in column_a:
                list = f' {list + str(cell.value)}\n'

            label_a.config(text=list)

        def get_b():
            list = ''
            for cell in column_b:
                list = f'{list + str(cell.value)}\n'

            label_b.config(text=list)

        Hardware_Button = Button(master, text="View Hardware", command=set_Sheet_Hardware)
        Hardware_Button.place(relx=.15, rely=.1, anchor=CENTER)
        Electrical_Button = Button(master, text="View Electrical", command=set_Sheet_Electrical)
        Electrical_Button.place(relx=.15, rely=.2, anchor=CENTER)
        Plumbing_Button = Button(master, text="View Plumbing", command=set_Sheet_plumbing)
        Plumbing_Button.place(relx=.15, rely=.3, anchor=CENTER)
        Flooring_Button = Button(master, text="View Flooring", command=set_Sheet_Flooring)
        Flooring_Button.place(relx=.15, rely=.4, anchor=CENTER)
        Lumber_Button = Button(master, text="View Lumber", command=set_Sheet_Lumber)
        Lumber_Button.place(relx=.15, rely=.5, anchor=CENTER)

        label_a = Label(master, text="")
        label_b = Label(master, text="")
        # Employee menu bar
        menubar = Menu(master)
        goHomeMenu = Menu(menubar, tearoff=0)
        goHomeMenu.add_command(label="Home", command=goHome)
        menubar.add_cascade(label="Home", menu=goHomeMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        master.config(menu=menubar)


class Activity_Log():
    def __init__(self, master, who):
        master.title("Activity Log")
        master.geometry("500x500")
        for widget in master.winfo_children():
            widget.destroy()

        def logOutBar():
            Login(root)

        def goHome():
            if who == "Admin":
                Admin(root)
            elif who == "Supervisor":
                Supervisor(root)
            else:
                Employee(root)

        for widget in master.winfo_children():
            widget.destroy()
        wb = Workbook()
        wb = load_workbook(activity_path)
        # Update Microsoft Excel Spreadsheet Function
        def updateExcel():
            # Get Size of Activity Log and start count at len(activity_size)
            wb = Workbook()
            wb = load_workbook(activity_path)

            wb.active = 0
            ws = (wb.active)
            column_b = ws['B']
            if ((employee.get() != "") & (description.get() != "") & (time.get() != "")):
                i = 2
                while (True):
                    cell1 = 'A' + str(i)
                    cell2 = 'B' + str(i)
                    cell3 = 'C' + str(i)
                    cell4 = 'D' + str(i)
                    if (ws[cell2].value == None):
                        ws[cell1] = i - 1
                        ws[cell2] = employee.get()
                        ws[cell3] = description.get()
                        ws[cell4] = time.get()
                        wb.save(activity_path)
                        test.grid_forget()
                        break
                    i += 1
                    wb.active = 0
                    ws = (wb.active)
                    global colB
                    colB = ws['B']
                    global colC
                    colC = ws['C']
                    global colD
                    colD = ws['D']
                    displayAct()

            else:
                test.grid(row=35, column=10, padx=10, pady=25)

        def displayAct():
            get_b()
            label_b.place(relx=.5, rely=.2, anchor=CENTER)
            get_c()
            label_c.place(relx=.7, rely=.2, anchor=CENTER)
            get_d()
            label_d.place(relx=.9, rely=.2, anchor=CENTER)


        def get_b():
            list = ''
            for cell in colB:
                list = f' {list + str(cell.value)}\n'

            label_b.config(text=list)

        def get_c():
            list = ''
            for cell in colC:
                list = f' {list + str(cell.value)}\n'

            label_c.config(text=list)

        def get_d():
            list = ''
            for cell in colD:
                list = f' {list + str(cell.value)}\n'

            label_d.config(text=list)


        label_b = Label(master, text="")
        label_c = Label(master, text="")
        label_d = Label(master, text="")
        #########
        global employee
        global description
        global time
        global test
        employee = StringVar()
        description = StringVar()
        time = StringVar()

        elabel = Label(master, text="Employee ID").grid(row=5, column=0, padx=10, pady=25)
        dlabel = Label(master, text="Description").grid(row=15, column=0, padx=10, pady=25)
        tlabel = Label(master, text="Time").grid(row=25, column=0, padx=10, pady=25)
        e = Entry(master, textvariable=employee, width=20).grid(row=5, column=5, padx=10, pady=25)
        d = Entry(master, textvariable=description, width=20).grid(row=15, column=5, padx=10, pady=25)
        t = Entry(master, textvariable=time, width=20).grid(row=25, column=5, padx=10, pady=25)

        update_button = Button(master, text="Update", command=updateExcel)
        update_button.grid(row=45, column=10, padx=10, pady=25)

        test = Label(master, text="All Sections must be filled out.", fg="red")
        test.grid(row=35, column=10, padx=10, pady=25)
        test.grid_forget()

        if who == "Admin":
            def delete():
                df = pd.DataFrame([[], [], []])
                with pd.ExcelWriter(activity_path) as writer:
                    df.to_excel(writer)
                deleted = Label(master, text="Activity Log Cleared").grid(row=1, column=5, padx=10, pady=25)

            delete_Buton = Button(master, text="Delete Activity Log", command=delete).grid(row=50, column=10, padx=10,
                                                                                           pady=25)

        # Employee menu bar
        menubar = Menu(master)
        goHomeMenu = Menu(menubar, tearoff=0)
        goHomeMenu.add_command(label="Home", command=goHome)
        menubar.add_cascade(label="Home", menu=goHomeMenu)
        logoutMenu = Menu(menubar, tearoff=0)
        logoutMenu.add_command(label="Log Out", command=logOutBar)
        menubar.add_cascade(label="Log Out", menu=logoutMenu)
        master.config(menu=menubar)


app = Login(root)
root.mainloop()
